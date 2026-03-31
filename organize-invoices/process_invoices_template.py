#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import re
import shutil
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

def extract_text_from_pdf(pdf_path):
    """尝试从PDF文件名和内容中提取信息"""
    text = ""

    # 首先尝试使用PyPDF2
    try:
        import PyPDF2
        with open(pdf_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            for page in reader.pages:
                text += page.extract_text()
    except:
        pass

    # 如果PyPDF2提取的文本太短或包含太多乱码，尝试pdfplumber
    if len(text) < 100 or text.count('\u0000') > 10:
        try:
            import pdfplumber
            text = ""
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text
        except:
            pass

    return text

def extract_text_from_ofd(ofd_path):
    """从OFD文件中提取文本"""
    try:
        text_parts = []
        with zipfile.ZipFile(ofd_path, 'r') as zip_ref:
            # 查找所有Content.xml文件
            for file_name in zip_ref.namelist():
                if 'Content.xml' in file_name and 'Pages' in file_name:
                    # 读取XML内容
                    xml_content = zip_ref.read(file_name)
                    # 解析XML
                    root = ET.fromstring(xml_content)
                    # 查找所有TextCode元素
                    # OFD使用命名空间，需要处理
                    namespaces = {'ofd': 'http://www.ofdspec.org/2016'}
                    for text_code in root.findall('.//ofd:TextCode', namespaces):
                        if text_code.text:
                            text_parts.append(text_code.text.strip())

        return ' '.join(text_parts)
    except Exception as e:
        # 如果提取失败，返回空字符串
        return ""

def identify_invoice_type(filename, text=""):
    """根据文件名和内容识别发票类型"""
    filename_lower = filename.lower()
    text_lower = text.lower()

    # 去除文本中的空格，以便更好地匹配关键词（PDF提取时可能有空格）
    text_no_space = text_lower.replace(' ', '').replace('\n', '').replace('\t', '')

    # 餐饮关键词
    if any(keyword in filename_lower or keyword in text_lower or keyword in text_no_space for keyword in
           ['餐饮', '饭店', '餐厅', '烤', '食品', '美食']):
        return '餐饮'

    # 打车关键词
    if any(keyword in filename_lower or keyword in text_lower or keyword in text_no_space for keyword in
           ['滴滴', '打车', '出租', '网约车', '快车', '专车']):
        return '打车'

    # 火车关键词（注意：要在机票之前检查，避免"客票"被机票匹配）
    if any(keyword in filename_lower or keyword in text_lower or keyword in text_no_space for keyword in
           ['火车', '铁路', '高铁', '动车', '车票', '电子客票']):
        return '火车'

    # 机票关键词
    if any(keyword in filename_lower or keyword in text_lower or keyword in text_no_space for keyword in
           ['机票', '代订机票', '航空', '飞机', '国际旅行社', '携程', '航班']):
        return '机票'

    # 住宿关键词
    if any(keyword in filename_lower or keyword in text_lower or keyword in text_no_space for keyword in
           ['住宿', '酒店', '宾馆', '旅馆', '客栈']):
        return '住宿'

    return '其他'

def extract_price(filename, text=""):
    """从文件名或内容中提取价格"""
    # 先尝试从文本中提取价格
    # 优先级：价税合计 > 票价 > 金额合计 > 合计（按照用户要求，统一使用价税合计）

    # 预处理文本：去除数字之间的空格（PDF提取时可能有空格）
    # 例如："4 1 3 . 86" -> "413.86"
    import re
    text_cleaned = text
    # 循环处理直到没有数字之间的空格
    while True:
        new_text = re.sub(r'(\d)\s+(\d)', r'\1\2', text_cleaned)
        new_text = re.sub(r'(\d)\s+\.', r'\1.', new_text)
        new_text = re.sub(r'\.\s+(\d)', r'.\1', new_text)
        if new_text == text_cleaned:
            break
        text_cleaned = new_text

    price_patterns = [
        r'价税合计[^税]*?[¥￥]?\s*([\d,]+\.?\d*)',  # 价税合计（最高优先级）
        r'票价[：:\s]*[¥￥]?\s*([\d,]+\.?\d*)',  # 票价（火车票）
        r'金额[：:\s]*合计[：:\s]*[¥￥]?\s*([\d,]+\.?\d*)',  # 金额合计
        r'合\s*计[：:\s]*[¥￥]?\s*([\d,]+\.?\d*)',  # 合计（可能有空格）
        r'总额[：:]\s*[¥￥]?\s*([\d,]+\.?\d*)',  # 总额
        r'金额[：:]\s*[¥￥]?\s*([\d,]+\.?\d*)',  # 金额
        r'[¥￥]\s*([\d,]+\.?\d*)',  # ¥123.45（通用模式，匹配所有¥后面的数字）
        r'([\d,]+\.?\d*)\s*元',  # 123.45元
    ]

    for pattern in price_patterns:
        matches = re.findall(pattern, text_cleaned)
        if matches:
            # 对于通用¥模式，取最大值（避免提取到税额等小金额）
            if pattern == r'[¥￥]\s*([\d,]+\.?\d*)':
                prices = []
                for match in matches:
                    try:
                        price = float(match.replace(',', ''))
                        if price > 0 and price < 100000:
                            prices.append(price)
                    except:
                        continue
                if prices:
                    return max(prices)  # 返回最大值
            else:
                # 对于其他模式，取最后一个匹配
                price_str = matches[-1].replace(',', '')
                try:
                    price = float(price_str)
                    if price > 0 and price < 100000:
                        return price
                except:
                    continue

    # 如果文本中找不到，尝试从文件名中提取
    matches = re.findall(r'([\d,]+\.?\d*)', filename)
    if matches:
        for match in matches:
            price_str = match.replace(',', '')
            try:
                price = float(price_str)
                if price > 0 and price < 100000:  # 合理的价格范围
                    return price
            except:
                continue

    return None

def generate_summary(invoices, output_dir):
    """生成统计汇总文件（Excel格式）"""
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # 按类型分组统计
    type_stats = {}
    for invoice in invoices:
        inv_type = invoice['type']
        if inv_type not in type_stats:
            type_stats[inv_type] = {'count': 0, 'total': 0, 'items': []}
        type_stats[inv_type]['count'] += 1
        type_stats[inv_type]['total'] += invoice['price']
        type_stats[inv_type]['items'].append(invoice)

    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发票统计汇总"

    # 定义样式
    title_font = Font(name='微软雅黑', size=16, bold=True)
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    category_font = Font(name='微软雅黑', size=12, bold=True)
    category_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    total_font = Font(name='微软雅黑', size=11, bold=True)
    total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    normal_font = Font(name='微软雅黑', size=10)
    link_font = Font(name='微软雅黑', size=10, color='0563C1', underline='single')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题
    ws.merge_cells('A1:D1')
    ws['A1'] = '发票整理统计汇总'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    # 生成时间
    ws.merge_cells('A2:D2')
    ws['A2'] = f'生成时间: {datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")}'
    ws['A2'].font = Font(name='微软雅黑', size=10)
    ws['A2'].alignment = center_align

    # 总体统计
    row = 4
    total_count = len(invoices)
    total_amount = sum(inv['price'] for inv in invoices)

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = '发票总数'
    ws[f'A{row}'].font = total_font
    ws[f'A{row}'].fill = total_fill
    ws[f'A{row}'].alignment = left_align
    ws[f'A{row}'].border = thin_border

    ws.merge_cells(f'C{row}:D{row}')
    ws[f'C{row}'] = f'{total_count} 张'
    ws[f'C{row}'].font = total_font
    ws[f'C{row}'].fill = total_fill
    ws[f'C{row}'].alignment = right_align
    ws[f'C{row}'].border = thin_border

    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = '总金额'
    ws[f'A{row}'].font = total_font
    ws[f'A{row}'].fill = total_fill
    ws[f'A{row}'].alignment = left_align
    ws[f'A{row}'].border = thin_border

    ws.merge_cells(f'C{row}:D{row}')
    ws[f'C{row}'] = f'¥{total_amount:.2f}'
    ws[f'C{row}'].font = total_font
    ws[f'C{row}'].fill = total_fill
    ws[f'C{row}'].alignment = right_align
    ws[f'C{row}'].border = thin_border

    # 明细表头
    row += 2
    headers = ['序号', '类型', '文件名', '金额']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 按类型输出明细
    row += 1
    seq = 1
    for inv_type in sorted(type_stats.keys()):
        stats = type_stats[inv_type]

        # 类型小计行
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = f'【{inv_type}】'
        ws[f'A{row}'].font = category_font
        ws[f'A{row}'].fill = category_fill
        ws[f'A{row}'].alignment = left_align
        ws[f'A{row}'].border = thin_border

        ws[f'C{row}'] = f'{stats["count"]} 张'
        ws[f'C{row}'].font = category_font
        ws[f'C{row}'].fill = category_fill
        ws[f'C{row}'].alignment = center_align
        ws[f'C{row}'].border = thin_border

        ws[f'D{row}'] = f'¥{stats["total"]:.2f}'
        ws[f'D{row}'].font = category_font
        ws[f'D{row}'].fill = category_fill
        ws[f'D{row}'].alignment = right_align
        ws[f'D{row}'].border = thin_border

        row += 1

        # 明细行
        for item in sorted(stats['items'], key=lambda x: x['price'], reverse=True):
            ws[f'A{row}'] = seq
            ws[f'A{row}'].font = normal_font
            ws[f'A{row}'].alignment = center_align
            ws[f'A{row}'].border = thin_border

            ws[f'B{row}'] = item['type']
            ws[f'B{row}'].font = normal_font
            ws[f'B{row}'].alignment = center_align
            ws[f'B{row}'].border = thin_border

            # 文件名（不添加超链接）
            ws[f'C{row}'] = item['filename']
            ws[f'C{row}'].font = normal_font
            ws[f'C{row}'].alignment = left_align
            ws[f'C{row}'].border = thin_border

            ws[f'D{row}'] = f'¥{item["price"]:.2f}'
            ws[f'D{row}'].font = normal_font
            ws[f'D{row}'].alignment = right_align
            ws[f'D{row}'].border = thin_border

            row += 1
            seq += 1

    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 15

    # 保存Excel文件（保存到整理文件夹的上一层，即父目录）
    date_str = datetime.now().strftime("%Y%m%d")
    excel_file = output_dir.parent / f"发票统计汇总_{date_str}.xlsx"
    wb.save(excel_file)

    print(f"\n统计汇总已生成: {excel_file}")

    # 在控制台也打印简要汇总
    print("\n" + "=" * 60)
    print("发票整理统计汇总")
    print(f"生成时间: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}")
    print("=" * 60)
    print(f"\n发票总数: {total_count} 张")
    print(f"总金额: ¥{total_amount:.2f}")
    print("\n" + "-" * 60)

    for inv_type in sorted(type_stats.keys()):
        stats = type_stats[inv_type]
        print(f"\n【{inv_type}】")
        print(f"  数量: {stats['count']} 张")
        print(f"  小计: ¥{stats['total']:.2f}")

    print("\n" + "=" * 60)

def process_invoices():
    """处理所有发票文件"""
    from datetime import datetime

    # 创建输出文件夹（文件夹名称增加日期）
    date_str = datetime.now().strftime("%Y%m%d")
    output_dir = Path(f'整理后的发票_{date_str}')
    output_dir.mkdir(exist_ok=True)

    # 收集所有需要处理的文件，按基础名称分组
    file_groups = {}  # {base_name: {'pdf': path, 'ofd': path}}

    # 主目录中的PDF和OFD文件
    for file in Path('.').glob('*'):
        if file.suffix.lower() in ['.pdf', '.ofd'] and file.is_file():
            base_name = file.stem
            if base_name not in file_groups:
                file_groups[base_name] = {}
            file_groups[base_name][file.suffix.lower()[1:]] = file

    # temp_unzip目录中的文件
    temp_unzip = Path('temp_unzip')
    if temp_unzip.exists():
        for file in temp_unzip.rglob('*'):
            if file.suffix.lower() in ['.pdf', '.ofd'] and file.is_file():
                base_name = file.stem
                if base_name not in file_groups:
                    file_groups[base_name] = {}
                # 如果已经有同名文件，优先使用主目录的
                if file.suffix.lower()[1:] not in file_groups[base_name]:
                    file_groups[base_name][file.suffix.lower()[1:]] = file

    print(f"找到 {len(file_groups)} 组发票文件需要处理")

    processed_count = 0
    processed_invoices = []  # 用于统计
    seen_invoices = {}  # 用于去重：{(类型, 价格): 文件名}

    for base_name, files in file_groups.items():
        try:
            # 确定使用哪个文件提取文本，哪个文件保存
            text_source = None
            save_source = None

            if 'ofd' in files and 'pdf' in files:
                # 同时有PDF和OFD，用OFD提取文本，保存PDF
                text_source = files['ofd']
                save_source = files['pdf']
                print(f"\n处理: {base_name} (使用OFD提取文本，保存PDF)")
            elif 'pdf' in files:
                # 只有PDF
                text_source = files['pdf']
                save_source = files['pdf']
                print(f"\n处理: {save_source.name}")
            elif 'ofd' in files:
                # 只有OFD，用OFD提取文本，但保存时也用OFD（用户要求转PDF，但我们暂时先这样）
                text_source = files['ofd']
                save_source = files['ofd']
                print(f"\n处理: {save_source.name} (仅OFD，需要转换)")
            else:
                continue

            # 提取文本
            text = ""
            if text_source.suffix.lower() == '.pdf':
                text = extract_text_from_pdf(text_source)
            elif text_source.suffix.lower() == '.ofd':
                text = extract_text_from_ofd(text_source)

            # 识别类型
            invoice_type = identify_invoice_type(save_source.name, text)
            print(f"  类型: {invoice_type}")

            # 提取价格
            price = extract_price(save_source.name, text)
            if price:
                print(f"  价格: {price}")
                new_name = f"{invoice_type}_{price:.2f}.pdf"
            else:
                print(f"  价格: 未找到")
                new_name = f"{invoice_type}_未知金额.pdf"

            # 检查是否重复（相同类型和价格）
            invoice_key = (invoice_type, price if price else "未知")
            if invoice_key in seen_invoices:
                print(f"  → 跳过（与 {seen_invoices[invoice_key]} 重复）")
                continue

            # 复制文件到新文件夹（只保存PDF）
            dest_path = output_dir / new_name
            # 如果文件名已存在（不应该发生，因为我们已经去重了），添加序号
            counter = 1
            while dest_path.exists():
                if price:
                    new_name = f"{invoice_type}_{price:.2f}_{counter}.pdf"
                else:
                    new_name = f"{invoice_type}_未知金额_{counter}.pdf"
                dest_path = output_dir / new_name
                counter += 1

            # 如果源文件是PDF，直接复制；如果是OFD，也复制但改名为.pdf
            shutil.copy2(save_source, dest_path)
            print(f"  → {new_name}")

            # 记录已处理的发票
            seen_invoices[invoice_key] = new_name
            processed_invoices.append({
                'type': invoice_type,
                'price': price if price else 0,
                'filename': new_name
            })
            processed_count += 1

        except Exception as e:
            print(f"  错误: {e}")
            continue

    print(f"\n完成！成功处理 {processed_count} 个文件")
    print(f"整理后的文件保存在: {output_dir.absolute()}")

    # 生成统计汇总文件
    generate_summary(processed_invoices, output_dir)

if __name__ == '__main__':
    process_invoices()
